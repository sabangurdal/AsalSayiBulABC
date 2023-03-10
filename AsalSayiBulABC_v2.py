import random
import os
from openpyxl import Workbook,load_workbook
from tqdm import tqdm
import time

class AsalSayiBulABC:
    def __init__(self,_Nektar_Sayisi,_Parametre_Sayisi,_Limit,_Alt_Sinir,_Ust_Sinir,_Dongu_Boyutu,_Evolüsyon_limiti):
        self.start=0
        self.durum=False
        self.Dongu=0;
        self.Dongu_Boyutu=_Dongu_Boyutu
        self.Nektar_Sayisi=_Nektar_Sayisi
        self.Nektarlar=[];
        self.Son_Limit=_Limit;
        self.parametre_sayisi=_Parametre_Sayisi;
        self.Alt_Sinir=_Alt_Sinir;
        self.Ust_Sinir=_Ust_Sinir;
        self.en_iyi_nektar=self.Nektar()
        self.toplam_fitness=0;
        self.bulundu=False;
        self.Evolüsyon__limiti=_Evolüsyon_limiti
        self.Evolüsyon_sayaci=0
        self.stop=0;
       
    class Nektar:
        def __init__(self):
            self.parametreler=[]
            self.Fitness=1;
            self.secilme_olasiligi=0;
            self.Limit=0;
            self.cesitlik=0;

    def Nektar_olustur(self):
        for x in range(self.Nektar_Sayisi):
            self.Nektarlar.append(self.Nektar())
            self.Nektarlar[x].parametreler=self.parametre_olustur()
            asal_sayi_miktarı=self.asal_sayisi_hesapla(self.Nektarlar[x]);
            self.Nektarlar[x].Fitness=self.fitness_hesapla(asal_sayi_miktarı);
            self.Nektarlar[x].cesitlik=self.cesitlik_hesapla(self.Nektarlar[x].parametreler)
        self.olasilik_hesaplama();

    def parametre_olustur(self):
        parametreler=[]
        olusan_parametre=0
        for x in range(self.parametre_sayisi):
            olusan_parametre=int(random.random()*(self.Ust_Sinir-self.Alt_Sinir)+self.Alt_Sinir)
            if(not(olusan_parametre==2)):
                while olusan_parametre%2==0:
                    olusan_parametre=int(random.random()*(self.Ust_Sinir-self.Alt_Sinir)+self.Alt_Sinir)
            parametreler.append(olusan_parametre)
        return parametreler;

    def cesitlik_hesapla(self,_parametreler):
        cesitlik=0
        for x in range(self.parametre_sayisi):
            tek=True
            for y in range(x+1,self.parametre_sayisi):
                if _parametreler[x]==_parametreler[y]:
                    tek=False
                    break
            if(tek and self.asal_kontrol(_parametreler[x])):
                cesitlik+=1
        return cesitlik  

    def asal_sayisi_hesapla(self,nektar):
        sayi=0;
        for x in range(self.parametre_sayisi):
            if(self.asal_kontrol(nektar.parametreler[x])):
                sayi+=1;
        return sayi;
        
    def fitness_hesapla(self,asal_sayi_miktarı):
        self.Evolüsyon_sayaci+=1
        return (1 / (asal_sayi_miktarı + 1))
                 
    def asal_kontrol(self, sayi):
        asalDurum=False;
        if sayi > 1:
           for i in range(2,sayi):
            if (sayi % i) == 0:
                asalDurum=False;
                break
            else:
                asalDurum=True;
        else:
            asalDurum=False;
        return asalDurum;

    def Isci_Ari_Fazi(self):
        i=0;#arı sayacı
        while(i<self.Nektar_Sayisi and self.Evolüsyon_sayaci<self.Evolüsyon__limiti):# Tüm Nektarlar İçin Komşu Nektar Üretimi
            _degisecekParametre_index=int(random.random()*self.parametre_sayisi)
            degisecekParametre=self.Nektarlar[i].parametreler[_degisecekParametre_index]
            while(self.asal_kontrol(degisecekParametre)):
                    if _degisecekParametre_index>0 :
                        degisecekParametre=self.Nektarlar[i].parametreler[_degisecekParametre_index-1]
                    if(self.asal_kontrol(degisecekParametre)):
                        if _degisecekParametre_index<self.parametre_sayisi-1: 
                            degisecekParametre=self.Nektarlar[i].parametreler[_degisecekParametre_index+1]
                            if(self.asal_kontrol(degisecekParametre)):
                                break
                    else:
                        break
                    _degisecekParametre_index=int(random.random()*self.parametre_sayisi)
                    degisecekParametre=self.Nektarlar[i].parametreler[_degisecekParametre_index]
            komsu_nektar_index=int(random.random()*self.Nektar_Sayisi)
            while(komsu_nektar_index==i):
                komsu_nektar_index=int(random.random()*self.Nektar_Sayisi)
            komsuParametre=self.Nektarlar[komsu_nektar_index].parametreler[_degisecekParametre_index]
            yeni_parametre_degeri=int(degisecekParametre+random.uniform(-1,1)*(degisecekParametre-komsuParametre))
            if(yeni_parametre_degeri>self.Ust_Sinir):
                yeni_parametre_degeri=self.Ust_Sinir
            if(yeni_parametre_degeri<self.Alt_Sinir):
                yeni_parametre_degeri=self.Alt_Sinir
            gecici_nektar=self.Nektar()
            gecici_nektar.parametreler=self.Nektarlar[i].parametreler.copy()
            gecici_nektar.parametreler[_degisecekParametre_index]=yeni_parametre_degeri;
            gecici_nektar.cesitlik=self.cesitlik_hesapla(gecici_nektar.parametreler)
            gecici_nektar.Fitness=self.fitness_hesapla(self.asal_sayisi_hesapla(gecici_nektar))
            if(self.Nektarlar[i].Fitness>gecici_nektar.Fitness):
                self.Nektarlar[i].parametreler=gecici_nektar.parametreler.copy()
                self.Nektarlar[i].cesitlik=gecici_nektar.cesitlik
                self.Nektarlar[i].Fitness=gecici_nektar.Fitness
            else:
                self.Nektarlar[i].Limit+=1 
            i=i+1
    
    def olasilik_hesaplama(self):
        self.toplam_fitness=0;
        for x in range(self.Nektar_Sayisi):
            self.toplam_fitness+=self.Nektarlar[x].Fitness  
        for y in range(self.Nektar_Sayisi):
            self.Nektarlar[y].secilme_olasiligi=self.Nektarlar[y].Fitness/self.toplam_fitness

    def Gozcu_Ari_Fazi(self):
        i=0;#besin sayacı
        t=0;#Arı Sayacı
        while(t<self.Nektar_Sayisi and self.Evolüsyon_sayaci<self.Evolüsyon__limiti):# Tüm Arılar İçin Olasılıksal seçim döngüsü
            r=random.random()
            if(r<self.Nektarlar[i].secilme_olasiligi):
                t+=1
                _degisecekParametre_index=int(random.random()*self.parametre_sayisi)
                degisecekParametre=self.Nektarlar[i].parametreler[_degisecekParametre_index]
                while(self.asal_kontrol(degisecekParametre)):
                    if _degisecekParametre_index>0 :
                        degisecekParametre=self.Nektarlar[i].parametreler[_degisecekParametre_index-1]
                    if(self.asal_kontrol(degisecekParametre)):
                        if _degisecekParametre_index<self.parametre_sayisi-1: 
                            degisecekParametre=self.Nektarlar[i].parametreler[_degisecekParametre_index+1]
                            if(self.asal_kontrol(degisecekParametre)):
                                break
                    else:
                        break
                    _degisecekParametre_index=int(random.random()*self.parametre_sayisi)
                    degisecekParametre=self.Nektarlar[i].parametreler[_degisecekParametre_index]
                komsu_nektar_index=int(random.random()*self.Nektar_Sayisi)
                while(komsu_nektar_index==i):
                    komsu_nektar_index=int(random.random()*self.Nektar_Sayisi)
                komsuParametre=self.Nektarlar[komsu_nektar_index].parametreler[_degisecekParametre_index]
                yeni_parametre_degeri=int(degisecekParametre+random.uniform(-1,1)*(degisecekParametre-komsuParametre))
                if(yeni_parametre_degeri>self.Ust_Sinir):
                    yeni_parametre_degeri=self.Ust_Sinir
                elif(yeni_parametre_degeri<self.Alt_Sinir):
                    yeni_parametre_degeri=self.Alt_Sinir
                gecici_nektar=self.Nektar()
                gecici_nektar.parametreler=self.Nektarlar[i].parametreler.copy()
                gecici_nektar.parametreler[_degisecekParametre_index]=yeni_parametre_degeri;
                gecici_nektar.cesitlik=self.cesitlik_hesapla(gecici_nektar.parametreler)
                gecici_nektar.Fitness=self.fitness_hesapla(self.asal_sayisi_hesapla(gecici_nektar))
                if(self.Nektarlar[i].Fitness>gecici_nektar.Fitness):
                    self.Nektarlar[i].parametreler=gecici_nektar.parametreler.copy()
                    self.Nektarlar[i].cesitlik=gecici_nektar.cesitlik
                    self.Nektarlar[i].Fitness=gecici_nektar.Fitness
                else:
                    self.Nektarlar[i].Limit+=1 
            i+=1
            i=i % self.Nektar_Sayisi #dağıtılmayan arılar için besin sayacını sıfırlıyoruz
    
    def En_iyi_belirleme(self):
        for i in range(self.Nektar_Sayisi):
            if(self.en_iyi_nektar.Fitness>self.Nektarlar[i].Fitness and 
               self.en_iyi_nektar.cesitlik<self.Nektarlar[i].cesitlik):
                self.en_iyi_nektar.parametreler=self.Nektarlar[i].parametreler.copy();
                self.en_iyi_nektar.cesitlik=self.Nektarlar[i].cesitlik
                self.en_iyi_nektar.Fitness=self.Nektarlar[i].Fitness

    def Kasif_Ari_Fazi(self):
        for i in range(self.Nektar_Sayisi):
            if(self.Nektarlar[i].Limit>=self.Son_Limit):
                self.Nektarlar[i].parametreler=self.parametre_olustur();
                self.Nektarlar[i].cesitlik=self.cesitlik_hesapla(self.Nektarlar[i].parametreler)
                self.Nektarlar[i].Limit=0
                self.Nektarlar[i].Fitness=self.fitness_hesapla(self.asal_sayisi_hesapla(self.Nektarlar[i]));

    def Dongu_Say(self):
        self.Dongu+=1
        if self.Dongu_Boyutu<=self.Dongu:
            self.durum=True
        if self.en_iyi_nektar.cesitlik==self.parametre_sayisi and self.en_iyi_nektar.Fitness<=1/self.parametre_sayisi+1:
            self.durum=True
            self.bulundu=True
        
    def rapor_ekle(self):
        eklenecek_dizi=[self.Dongu,self.Evolüsyon_sayaci,self.durum,self.stop,self.en_iyi_nektar.cesitlik]
        for x in range(self.parametre_sayisi):
            eklenecek_dizi.append(self.en_iyi_nektar.parametreler[x])
        wb = load_workbook("data.xlsx")
        ws = wb["Algoritma 2"]
        ws.append(eklenecek_dizi)
        wb.save("data.xlsx")
    
    def rapor_olustur(self):
       eklenecek_dizi=["Döngü","Evolüsyon","Durum","Süre","Cesitlik"]
       for x in range(self.parametre_sayisi):
           eklenecek_dizi.append("Sayi "+str(x))
       wb = Workbook()
       ws = wb.active
       ws.title = "Algoritma 1"
       ws.append(eklenecek_dizi)
       ws = wb.create_sheet("Algoritma 2")
       ws.append(eklenecek_dizi)
       wb.save("data.xlsx")
 
    def main(self):
        self.start=time.time()
        self.Nektar_olustur();
        while(not(self.durum) and self.Evolüsyon_sayaci<self.Evolüsyon__limiti):
            self.Isci_Ari_Fazi();
            self.olasilik_hesaplama();
            self.Gozcu_Ari_Fazi();
            self.Kasif_Ari_Fazi();
            self.En_iyi_belirleme();
            self.Dongu_Say();
        if(self.bulundu):
           print(self.Dongu,"Döngüde",self.Evolüsyon_sayaci," Evolüsyonda tüm Kısıtlar Sağlandı.Bulunan Çözüm:")
           print("Parametreler:",self.en_iyi_nektar.parametreler,"Çeşitlilik:",self.en_iyi_nektar.cesitlik)
        else:
            print("Tüm Kısıtlar",self.Dongu, " Döngüde ve ",self.Evolüsyon_sayaci," Evolüsyonda Sağlanamadı.Bulunan En İyi Çözüm:")
            print("Parametreler:",self.en_iyi_nektar.parametreler,"Çeşitlilik:",self.en_iyi_nektar.cesitlik)
        self.stop=time.time()-self.start;
        if(not (os.path.exists("data.xlsx"))):
            self.rapor_olustur();
            self.rapor_ekle();
        else:
            self.rapor_ekle()

    
dongu_sayisi=5
nektar_Miktarı=10
parametre_sayisi=20
limit=1000
alt=2
ust=2000
dongu=1000
Evolüsyon_Limiti=4000;

for x in range(100):
    ABC=AsalSayiBulABC(nektar_Miktarı,parametre_sayisi,limit,alt,ust,dongu,Evolüsyon_Limiti);
    ABC.main();


